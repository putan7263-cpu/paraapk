# -*- coding: utf-8 -*-
"""Генерирует Наценки.xlsx с расчётами по 2 регионам.
Запуск:  .venv\\Scripts\\python.exe make_excel.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from api.constants import REGION_BRACKETS


REGION_INFO = {
    "en-tr": {
        "title":         "Турция",
        "currency":      "₺",
        "rate":          1.80,   # ₽ за 1 ₺ для обычных цен
        "low_rate":      3.20,   # ₽ за 1 ₺ для дешёвых игр (≤ порога)
        "low_threshold": 150,    # порог в локальной валюте
    },
    "ru-ua": {
        "title":         "Украина",
        "currency":      "₴",
        "rate":          1.88,   # ₽ за 1 ₴
        "low_rate":      None,   # нет двухтарифной модели
        "low_threshold": None,
    },
}

OUT_FILE = "Наценки.xlsx"


# ─── стили ────────────────────────────────────────────────────────────────

THIN     = Side(border_style="thin",   color="BFBFBF")
THICK    = Side(border_style="medium", color="2F5496")
ALL_BRD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
F_TITLE  = Font(bold=True, size=16, color="2F5496")
F_SECT   = Font(bold=True, size=12, color="FFFFFF")
F_HDR    = Font(bold=True, size=11, color="FFFFFF")
F_BOLD   = Font(bold=True)
F_PRICE  = Font(bold=True, size=12, color="2F5496")
FILL_SECT   = PatternFill("solid", fgColor="2F5496")
FILL_HDR    = PatternFill("solid", fgColor="4472C4")
FILL_EDIT   = PatternFill("solid", fgColor="FFF2CC")    # жёлтый — редактируемое
FILL_RESULT = PatternFill("solid", fgColor="E2EFDA")    # зелёноватый — итоги
FILL_ALT    = PatternFill("solid", fgColor="F2F2F2")    # серая полоска
CENTER  = Alignment(horizontal="center", vertical="center")
LEFT    = Alignment(horizontal="left",   vertical="center", indent=1)


# ─── построение листа региона ─────────────────────────────────────────────

def build_sheet(ws, code: str, info: dict):
    sym  = info["currency"]
    rate = info["rate"]
    low_rate      = info.get("low_rate")
    low_threshold = info.get("low_threshold")
    has_tier      = low_rate is not None and low_threshold is not None
    brackets = REGION_BRACKETS[code]

    # ── заголовок ────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    ws["A1"] = f"Наценки и расчёты — {info['title']} ({code})"
    ws["A1"].font = F_TITLE
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 26

    # ── курс (редактируемые ячейки) ──────────────────────────────────────
    label_rate = f"Курс ₽ за 1 {sym} (обычный):" if has_tier else f"Курс ₽ за 1 {sym}:"
    ws["A3"] = label_rate
    ws["A3"].font = F_BOLD
    ws["A3"].alignment = LEFT
    ws["B3"] = rate
    ws["B3"].font = F_BOLD
    ws["B3"].fill = FILL_EDIT
    ws["B3"].border = ALL_BRD
    ws["B3"].alignment = CENTER
    ws["B3"].number_format = "0.0000"
    ws["C3"] = "← редактируемое (жёлтое)"
    ws["C3"].font = Font(italic=True, color="7F7F7F", size=10)

    RATE     = "$B$3"
    LOW_RATE = None
    LOW_THR  = None

    if has_tier:
        ws["A4"] = f"Курс ₽ за 1 {sym} (на дешёвые игры):"
        ws["A4"].font = F_BOLD
        ws["A4"].alignment = LEFT
        ws["B4"] = low_rate
        ws["B4"].font = F_BOLD
        ws["B4"].fill = FILL_EDIT
        ws["B4"].border = ALL_BRD
        ws["B4"].alignment = CENTER
        ws["B4"].number_format = "0.0000"

        ws["A5"] = f"Порог дешёвой цены {sym} (≤):"
        ws["A5"].font = F_BOLD
        ws["A5"].alignment = LEFT
        ws["B5"] = low_threshold
        ws["B5"].font = F_BOLD
        ws["B5"].fill = FILL_EDIT
        ws["B5"].border = ALL_BRD
        ws["B5"].alignment = CENTER
        ws["B5"].number_format = "0"
        ws["C5"] = f"при цене ≤ {sym} {low_threshold} используется второй курс"
        ws["C5"].font = Font(italic=True, color="7F7F7F", size=10)

        LOW_RATE = "$B$4"
        LOW_THR  = "$B$5"

    # вспомогательная формула «эффективный курс для цены X»
    def eff_rate(price_ref: str) -> str:
        if has_tier:
            return f"IF({price_ref}<={LOW_THR},{LOW_RATE},{RATE})"
        return RATE

    # ── секция «Калькулятор» ─────────────────────────────────────────────
    section_row = 7 if has_tier else 5
    ws.merge_cells(f"A{section_row}:B{section_row}")
    sect_cell = ws.cell(row=section_row, column=1, value="  КАЛЬКУЛЯТОР ЦЕНЫ")
    sect_cell.font = F_SECT
    sect_cell.fill = FILL_SECT
    sect_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[section_row].height = 22

    price_row = section_row + 1   # ввод цены
    # на сколько строк калькулятор: 8 (с курсом-индикатором) или 7
    calc_rows = 8 if has_tier else 7
    TABLE_HEADER_ROW = price_row + calc_rows + 1
    TABLE_START = TABLE_HEADER_ROW + 1
    TABLE_END   = TABLE_START + len(brackets) - 1

    lookup_range = f"$A${TABLE_START}:$D${TABLE_END}"

    # пользовательский ввод цены
    ws.cell(row=price_row, column=1, value=f"Введите цену {sym}:").font = F_BOLD
    ws.cell(row=price_row, column=1).alignment = LEFT
    pcell = ws.cell(row=price_row, column=2, value=1000)
    pcell.fill = FILL_EDIT
    pcell.border = ALL_BRD
    pcell.alignment = CENTER
    pcell.number_format = f"#,##0.00 \"{sym}\""

    PRICE = f"B{price_row}"

    # производные строки
    def calc_row(row, label, formula, fmt, *, bold=False, fill=None, comment=None):
        a = ws.cell(row=row, column=1, value=label)
        a.alignment = LEFT
        b = ws.cell(row=row, column=2, value=formula)
        b.number_format = fmt
        b.border = ALL_BRD
        b.alignment = CENTER
        if bold:
            a.font = F_BOLD
            b.font = F_PRICE
        if fill is not None:
            b.fill = fill
        if comment:
            c = ws.cell(row=row, column=3, value=comment)
            c.font = Font(italic=True, color="7F7F7F", size=10)

    r = price_row + 1
    if has_tier:
        calc_row(r, "Применённый курс ₽:", f"={eff_rate(PRICE)}", "0.0000",
                 comment=f"= IF(цена ≤ B5, B4, B3)")
        r += 1
        # закуп: считаем напрямую от цены через efficient rate
        calc_row(r, "Закуп (себестоимость) ₽:", f"={PRICE}*({eff_rate(PRICE)})", "#,##0.00 \"₽\"")
    else:
        calc_row(r, "Закуп (себестоимость) ₽:", f"={PRICE}*{RATE}", "#,##0.00 \"₽\"")
    cost_row = r
    r += 1
    calc_row(r, "Применённый коэффициент:", f"=VLOOKUP({PRICE},{lookup_range},3,TRUE)", "0.00")
    coef_row = r
    r += 1
    calc_row(r, "Применённая добавка ₽:",   f"=VLOOKUP({PRICE},{lookup_range},4,TRUE)", "#,##0 \"₽\"")
    add_row = r
    r += 1
    calc_row(r, "ПРОДАЖА ₽ (твоя цена):",
             f"=MROUND({PRICE}*B{coef_row}+B{add_row},10)",
             "#,##0 \"₽\"", bold=True, fill=FILL_RESULT)
    sell_row = r
    r += 1
    calc_row(r, "Прибыль ₽:", f"=B{sell_row}-B{cost_row}", "#,##0.00 \"₽\"")
    profit_row = r
    r += 1
    calc_row(r, "Наценка % (к закупу):", f"=IFERROR(B{profit_row}/B{cost_row},0)", "0.0 %")
    r += 1
    calc_row(r, "Маржа % (от продажи):", f"=IFERROR(B{profit_row}/B{sell_row},0)", "0.0 %")

    # ── секция «Таблица коэффициентов» ───────────────────────────────────
    headers = [
        f"Мин {sym}", f"Макс {sym}",
        "Коэф", "Доб ₽",
        f"Расч.цена {sym}", "Закуп ₽",
        "Продажа ₽", "Прибыль ₽",
        "Наценка %", "Маржа %",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=TABLE_HEADER_ROW, column=col, value=h)
        c.font = F_HDR
        c.fill = FILL_HDR
        c.alignment = CENTER
        c.border = ALL_BRD
    ws.row_dimensions[TABLE_HEADER_ROW].height = 24

    # ── данные таблицы ───────────────────────────────────────────────────
    for i, (mn, mx, coef, add) in enumerate(brackets):
        r = TABLE_START + i
        fill = FILL_ALT if i % 2 else None
        # Мин, Макс, Коэф, Доб
        for col, val, nf in [
            (1, int(mn),   "0"),
            (2, int(mx),   "0"),
            (3, float(coef), "0.00"),
            (4, float(add),  "0"),
        ]:
            c = ws.cell(row=r, column=col, value=val)
            c.alignment = CENTER
            c.border = ALL_BRD
            c.number_format = nf
            if col == 3:
                c.font = F_BOLD
            if fill:
                c.fill = fill

        # E — Расчётная цена = середина диапазона
        e = ws.cell(row=r, column=5, value=f"=(A{r}+B{r})/2")
        e.number_format = f"#,##0 \"{sym}\""
        e.alignment = CENTER
        e.border = ALL_BRD
        if fill: e.fill = fill

        # F — Закуп ₽ = E * курс (двухтарифный для Турции)
        if has_tier:
            cost_formula = f"=E{r}*IF(E{r}<={LOW_THR},{LOW_RATE},{RATE})"
        else:
            cost_formula = f"=E{r}*{RATE}"
        f = ws.cell(row=r, column=6, value=cost_formula)
        f.number_format = "#,##0 \"₽\""
        f.alignment = CENTER
        f.border = ALL_BRD
        if fill: f.fill = fill

        # G — Продажа ₽ = MROUND(E*Coef + Add, 10)
        g = ws.cell(row=r, column=7, value=f"=MROUND(E{r}*C{r}+D{r},10)")
        g.number_format = "#,##0 \"₽\""
        g.alignment = CENTER
        g.border = ALL_BRD
        g.font = F_BOLD
        if fill: g.fill = fill

        # H — Прибыль = G - F
        h = ws.cell(row=r, column=8, value=f"=G{r}-F{r}")
        h.number_format = "#,##0 \"₽\""
        h.alignment = CENTER
        h.border = ALL_BRD
        if fill: h.fill = fill

        # I — Наценка % = (G-F)/F
        i_cell = ws.cell(row=r, column=9, value=f"=IFERROR((G{r}-F{r})/F{r},0)")
        i_cell.number_format = "0 %"
        i_cell.alignment = CENTER
        i_cell.border = ALL_BRD
        if fill: i_cell.fill = fill

        # J — Маржа % = (G-F)/G
        j_cell = ws.cell(row=r, column=10, value=f"=IFERROR((G{r}-F{r})/G{r},0)")
        j_cell.number_format = "0 %"
        j_cell.alignment = CENTER
        j_cell.border = ALL_BRD
        if fill: j_cell.fill = fill

    # ── условное форматирование на колонку «Наценка %» ──────────────────
    rule = ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B",
    )
    ws.conditional_formatting.add(f"I{TABLE_START}:I{TABLE_END}", rule)

    # ── ширина колонок ───────────────────────────────────────────────────
    widths = {
        "A": 28, "B": 16, "C": 16, "D": 12, "E": 16,
        "F": 14, "G": 14, "H": 14, "I": 12, "J": 12,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── закрепляем шапку таблицы ─────────────────────────────────────────
    ws.freeze_panes = f"A{TABLE_START}"

    # ── легенда внизу ────────────────────────────────────────────────────
    legend_r = TABLE_END + 2
    ws.cell(row=legend_r, column=1, value="Легенда:").font = F_BOLD

    notes = [
        ("Жёлтые ячейки",   "редактируемые (курсы, порог, цена калькулятора)"),
        ("Зелёная ячейка",  "итоговая твоя цена продажи в ₽"),
        ("Расч. цена",      "середина диапазона — взята для наглядного расчёта"),
        ("Продажа",         "формула как в программе: MROUND(цена×коэф + добавка, 10)"),
        ("Наценка %",       "(Продажа − Закуп) / Закуп — сколько накручиваешь сверху"),
        ("Маржа %",         "(Продажа − Закуп) / Продажа — какую долю составляет твоя прибыль"),
    ]
    if has_tier:
        notes.append((
            "Двухтарифный курс",
            f"если цена ≤ {sym} {low_threshold} — закуп считается по курсу B4 ({low_rate} ₽/{sym}), иначе по B3",
        ))
    for k, (n, d) in enumerate(notes):
        ws.cell(row=legend_r + 1 + k, column=1, value=n).font = Font(italic=True, color="2F5496")
        ws.cell(row=legend_r + 1 + k, column=2, value=d).alignment = LEFT


# ─── собираем книгу ───────────────────────────────────────────────────────

def main():
    wb = Workbook()
    wb.remove(wb.active)

    for code, info in REGION_INFO.items():
        ws = wb.create_sheet(title=f"{info['title']} ({code})")
        build_sheet(ws, code, info)

    wb.save(OUT_FILE)
    print(f"Готово: {OUT_FILE}")


if __name__ == "__main__":
    main()
